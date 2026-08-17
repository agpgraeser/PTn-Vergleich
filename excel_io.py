"""
excel_io.py – Einlesen der Messdaten-Arbeitsmappen.

Portiert aus `src/utils/excelParser.ts` der React-App (2026-08-17), damit
dieselben Dateien wie bisher gelesen werden koennen.

Erkannt werden zwei Aufbauten:

Aufbau C – drei Kopfzeilen:
    Zeile 1: Systemname
    Zeile 2: Variablennamen  (Zeit | y | u)
    Zeile 3: Einheiten
    ab Zeile 4: Messwerte

Aufbau A/B – keine oder eine Kopfzeile, Spalten Zeit | y | u.
"""

from __future__ import annotations

import io

from openpyxl import load_workbook


class ExcelError(ValueError):
    """Fehler mit verstaendlicher Meldung fuer das Frontend."""


def _zahl(wert) -> float | None:
    """Wandelt eine Zelle in eine Zahl; gibt None zurueck wenn das nicht geht."""
    if wert is None or isinstance(wert, bool):
        return None
    if isinstance(wert, (int, float)):
        return float(wert)
    text = str(wert).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def messdaten_lesen(inhalt: bytes, dateiname: str = "") -> dict:
    """Liest die erste Tabelle der Arbeitsmappe.

    Rueckgabe: zeit / y_daten / u_daten (Listen), systemname, var_namen,
    einheiten und die ersten drei Zeilen als Vorschau.
    """
    try:
        wb = load_workbook(io.BytesIO(inhalt), data_only=True, read_only=True)
    except Exception as ex:
        raise ExcelError(f"Datei konnte nicht gelesen werden: {ex}")

    try:
        ws = wb[wb.sheetnames[0]]
        zeilen = [list(z) for z in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not zeilen:
        raise ExcelError("Die Tabelle ist leer.")

    basisname = dateiname.rsplit(".", 1)[0] if "." in dateiname else dateiname

    erste = zeilen[0] if len(zeilen) > 0 else []
    zweite = zeilen[1] if len(zeilen) > 1 else []
    dritte = zeilen[2] if len(zeilen) > 2 else []

    def ist_text(zeile, spalte=0):
        return (len(zeile) > spalte
                and isinstance(zeile[spalte], str)
                and zeile[spalte].strip() != "")

    # Aufbau C: Kopfzeile 1 und 2 sind Text, es gibt eine dritte Zeile
    aufbau_c = ist_text(erste) and ist_text(zweite) and len(dritte) > 0

    def txt(zeile, i, vorgabe):
        if len(zeile) > i and zeile[i] not in (None, ""):
            return str(zeile[i])
        return vorgabe

    if aufbau_c:
        systemname = str(erste[0])
        var_namen = {"zeit": txt(zweite, 0, "Zeit"),
                     "y": txt(zweite, 1, "y"),
                     "u": txt(zweite, 2, "u")}
        einheiten = {"zeit": txt(dritte, 0, ""),
                     "y": txt(dritte, 1, ""),
                     "u": txt(dritte, 2, "")}
        daten_ab = 3
        hat_u = len(zweite) >= 3 and zweite[2] not in (None, "")
    else:
        systemname = basisname
        var_namen = {"zeit": "Zeit", "y": "y", "u": "u"}
        einheiten = {"zeit": "", "y": "", "u": ""}
        daten_ab = 0
        hat_u = len(erste) >= 3

    zeit: list[float] = []
    y_daten: list[float] = []
    u_daten: list[float] = []

    for zeile in zeilen[daten_ab:]:
        if not zeile or len(zeile) < 2:
            continue
        t = _zahl(zeile[0])
        y = _zahl(zeile[1])
        if t is None or y is None:
            continue                      # Kopf-/Leerzeilen ueberspringen
        zeit.append(t)
        y_daten.append(y)
        if hat_u and len(zeile) >= 3:
            u = _zahl(zeile[2])
            u_daten.append(u if u is not None else 0.0)

    if not zeit:
        raise ExcelError("Keine auswertbaren Messwerte gefunden "
                         "(erwartet: Spalte 1 Zeit, Spalte 2 Messgroesse).")

    vorschau = [[("" if z is None else str(z)) for z in (zeile or [])[:3]]
                for zeile in zeilen[:3]]

    return {
        "zeit": zeit,
        "y_daten": y_daten,
        "u_daten": u_daten,
        "systemname": systemname,
        "var_namen": var_namen,
        "einheiten": einheiten,
        "vorschau": vorschau,
        "dateiname": dateiname,
    }


def messung_aus_zeitverlaeufen(zeitverlaeufe: dict, dateiname: str = "") -> dict | None:
    """Formt das Blatt `Zeitverlaeufe` einer AGP-Projektdatei in dasselbe
    Format wie `messdaten_lesen`, damit die Oberflaeche beide Quellen
    gleich behandeln kann.

    Erwartet Spalten t / y / u (Gross-/Kleinschreibung egal). y ist die
    Regelgroesse, u die Stellgroesse.
    """
    # Achtung: in der Projektdatei ist `daten` SPALTENweise abgelegt –
    # daten[i] ist die komplette Spalte i, nicht eine Zeile.
    spalten = [str(s) for s in (zeitverlaeufe.get("spalten") or [])]
    daten = zeitverlaeufe.get("daten") or []
    if not spalten or not daten:
        return None

    def index_von(*kandidaten):
        for k in kandidaten:
            for i, s in enumerate(spalten):
                if s.strip().lower() == k:
                    return i
        return None

    i_t = index_von("t", "zeit", "time")
    i_y = index_von("y", "regelgroesse", "regelgröße")
    i_u = index_von("u", "stellgroesse", "stellgröße")
    if i_t is None or i_y is None:
        return None
    if max(i_t, i_y) >= len(daten):
        return None

    spalte_t = daten[i_t]
    spalte_y = daten[i_y]
    spalte_u = daten[i_u] if (i_u is not None and i_u < len(daten)) else None

    zeit, y_daten, u_daten = [], [], []
    for k in range(min(len(spalte_t), len(spalte_y))):
        t = _zahl(spalte_t[k])
        y = _zahl(spalte_y[k])
        if t is None or y is None:
            continue
        zeit.append(t)
        y_daten.append(y)
        if spalte_u is not None and k < len(spalte_u):
            u = _zahl(spalte_u[k])
            u_daten.append(u if u is not None else 0.0)

    if not zeit:
        return None

    return {
        "zeit": zeit,
        "y_daten": y_daten,
        "u_daten": u_daten,
        "systemname": "",
        "var_namen": {"zeit": "Zeit", "y": "y", "u": "u"},
        "einheiten": {"zeit": "", "y": "", "u": ""},
        "vorschau": [],
        "dateiname": dateiname,
    }
